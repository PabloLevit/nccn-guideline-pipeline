from __future__ import annotations

import difflib
import json
import re
import time
import argparse
import os
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Set

import requests
from bs4 import BeautifulSoup

import openpyxl
import pdfplumber

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# ============================================================
# NOISE FILTER (EULA / headers / footer junk)
# ============================================================
NOISE_LINE_RE = re.compile(
    r"(PLEASE NOTE|End-User License Agreement|MAY NOT distribute|"
    r"Copyright|All Rights Reserved|NCCN Guidelines Index|Table of Contents|"
    r"Printed by|may not be reproduced|NCCN Guidelines Version)",
    re.IGNORECASE
)

def _is_noise_line(s: str) -> bool:
    return bool(NOISE_LINE_RE.search((s or "").strip()))

def _strip_control_chars(s: str) -> str:
    # remove weird PDF control chars (e.g., \x0c)
    return re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", s or "")


# ============================================================
# USER PATHS (portable defaults; can be overridden via CLI or env vars)
# ============================================================
PROJECT_DIR = Path(__file__).resolve().parent

PDF_GUIDELINES_DIR = Path(
    os.getenv("PDF_GUIDELINES_DIR", str(PROJECT_DIR / "PDF Guidelines"))
).resolve()

BASE_DIR = PROJECT_DIR

SECRETS_PATH = Path(
    os.getenv("SECRETS_PATH", str(BASE_DIR / "secrets.json"))
).resolve()

EXCEL_PATH_DEFAULT = Path(
    os.getenv("EXCEL_PATH", str(BASE_DIR / "data" / "NCCN_NGS_recommendations_2025_SANDBOX.xlsx"))
).resolve()

DOWNLOAD_DIR_DEFAULT = Path(
    os.getenv("DOWNLOAD_DIR", str(BASE_DIR / "downloads"))
).resolve()

OUT_DIR_DEFAULT = Path(
    os.getenv("OUT_DIR", str(BASE_DIR / "update_reports"))
).resolve()

# Dedicated Selenium profile (optional; keep env-driven)
CHROME_USER_DATA_DIR = os.getenv(
    "CHROME_USER_DATA_DIR",
    str((BASE_DIR / "chrome_profile").resolve())
)
CHROME_PROFILE_DIR = os.getenv("CHROME_PROFILE_DIR", "Default")


# ============================================================
# NCCN URLs
# ============================================================
BASE_URL = "https://www.nccn.org"
CATEGORY_1_URL = f"{BASE_URL}/guidelines/category_1"
HEADERS = {"User-Agent": "Mozilla/5.0"}

# ============================================================
# 11 Questions (anchors)
# ============================================================
QUESTION_ANCHORS: List[Tuple[str, str]] = [
    ("1.1.1", "Is there a recommendation for Tissue NGS?"),
    ("1.1.2", "For which disease stage(s) does the guideline recommend NGS or molecular testing?"),
    ("1.2",   "What are the key biomarkers?"),
    ("1.3",   "Is there a recommendation for PDL1 / MSI / TMB test?"),
    ("2.",    "Is there a recommendation for ctDNA NGS?"),
    ("3.1",   "Is there a recommendation for RNA sequencing?"),
    ("3.2",   "What are the key biomarkers mentioned for the RNA testing?"),
    ("4.",    "Is there a recommendation to use a CLIA-approved assay/lab?"),
    ("5.",    "Is there a recommendation for MRD (Minimal Residual Disease) testing?"),
    ("6.",    "Is there any mention of the CTC (Circulating Tumor Cells) test?"),
    ("7.",    "Is there any mention of the Chemosensitivity or Functional test?"),
]

Q_MAP: Dict[str, List[str]] = {
    "1.1.1": ["ngs", "molecular", "sequencing", "molecular profiling", "broad panel", "biomarker testing"],
    "1.1.2": ["stage", "resectable", "unresectable", "metastatic", "advanced", "adjuvant", "neoadjuvant", "surveillance"],
    "1.2":   ["biomarker", "egfr", "alk", "ros1", "ret", "ntrk", "kras", "erbb2", "her2", "met", "braf", "msi", "tmb", "pd-l1", "cldn18.2"],
    "1.3":   ["pd-l1", "pdl1", "msi", "tmb", "cps", "tap"],
    "2.":    ["ctdna", "plasma", "liquid biopsy", "cfdna"],
    "3.1":   ["rna", "rna-based", "transcript", "fusion", "rearrangement"],
    "3.2":   ["rna", "fusion", "rearrangement", "nrg1", "ntrk", "ret", "ros1", "alk"],
    "4.":    ["clia", "fda-approved", "assay", "laboratory", "companion diagnostic"],
    "5.":    ["mrd", "minimal residual disease"],
    "6.":    ["ctc", "circulating tumor cells"],
    "7.":    ["chemosensitivity", "functional", "ex vivo", "drug sensitivity", "resistance assay", "organoid"],
}

VERSION_RE = re.compile(r"\bVersion[:\s]*([\d.]+)\b", re.IGNORECASE)

SECTION_RE = re.compile(
    r"\b([A-Z]{2,12}-[A-Z0-9]{1,12})"
    r"(?:\s*,?\s*(\d+(?:-\d+)?\s*of\s*\d+))?"
    r"(?:\s*,?\s*Page\s*(\d+(?:-\d+)?))?\b"
)

MS_RE = re.compile(r"\bMS-\d+\b")

# A “real” NCCN section header like: GAST-B 3 of 7
SECTION_HEADER_LINE_RE = re.compile(
    r"^\s*([A-Z]{2,12}-[A-Z0-9]{1,12})\s+(\d+(?:-\d+)?)\s+of\s+(\d+)\s*$"
)

# Sometimes NCCN gives "GAST-B 3/7" or variants; keep a softer fallback
SECTION_HEADER_SOFT_RE = re.compile(
    r"^\s*([A-Z]{2,12}-[A-Z0-9]{1,12})\s+(\d+(?:-\d+)?)\s*(?:of|/)\s*(\d+)\s*$",
    re.IGNORECASE
)


# ============================================================
# Data structures
# ============================================================
@dataclass
class QuestionBlock:
    qid: str
    anchor_row: int
    start_row: int
    end_row: int


# ============================================================
# Small utilities
# ============================================================
def ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)

def load_json(path: Path) -> dict:
    if not path.exists():
        raise FileNotFoundError(f"Missing file: {path}")
    return json.loads(path.read_text(encoding="utf-8"))

def cell_text(v) -> str:
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    return str(v).strip()

def slugify(s: str) -> str:
    s = s.strip()
    s = re.sub(r"[^\w\s-]", "", s)
    s = re.sub(r"\s+", "_", s)
    return s[:120] if len(s) > 120 else s

def normalize_whitespace(s: str) -> str:
    # IMPORTANT: keep this for Excel lines (single-line), NOT for Updates blocks.
    s = s.replace("\n", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


# ============================================================
# Excel parsing
# ============================================================
def find_reference_column(ws) -> Optional[int]:
    for row in range(1, min(30, ws.max_row) + 1):
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row=row, column=col).value
            if isinstance(v, str) and v.strip().lower() == "references":
                return col
    return None

def find_question_rows(ws) -> Dict[str, int]:
    found: Dict[str, int] = {}
    max_scan_cols = min(12, ws.max_column)

    for row in range(1, ws.max_row + 1):
        row_text = " | ".join(cell_text(ws.cell(row=row, column=c).value) for c in range(1, max_scan_cols + 1))
        low = row_text.lower()
        for qid, qtxt in QUESTION_ANCHORS:
            if qid in found:
                continue
            if qid.lower() in low and qtxt.lower() in low:
                found[qid] = row
    return found

def build_blocks(question_rows: Dict[str, int], ws_max_row: int) -> List[QuestionBlock]:
    items = sorted(question_rows.items(), key=lambda x: x[1])
    blocks: List[QuestionBlock] = []
    for i, (qid, anchor_row) in enumerate(items):
        start_row = anchor_row + 1
        end_row = (items[i + 1][1] - 1) if i + 1 < len(items) else ws_max_row
        blocks.append(QuestionBlock(qid=qid, anchor_row=anchor_row, start_row=start_row, end_row=end_row))
    return blocks

def extract_existing_block_text(ws, block: QuestionBlock, ref_col: Optional[int]) -> Tuple[List[str], List[str]]:
    answers: List[str] = []
    refs: List[str] = []

    for r in range(block.start_row, block.end_row + 1):
        row_vals: List[str] = []
        for c in range(1, ws.max_column + 1):
            if ref_col and c == ref_col:
                continue
            txt = cell_text(ws.cell(row=r, column=c).value)
            if txt:
                row_vals.append(txt)
        if row_vals:
            answers.append(" ".join(row_vals))

        if ref_col:
            ref_txt = cell_text(ws.cell(row=r, column=ref_col).value)
            if ref_txt:
                refs.append(ref_txt)

    def dedup(seq: List[str]) -> List[str]:
        seen = set()
        out = []
        for x in seq:
            if x not in seen:
                out.append(x)
                seen.add(x)
        return out

    return dedup(answers), dedup(refs)

def find_best_sheet(wb, tumor_input: str, guideline_title: str) -> Optional[str]:
    candidates = wb.sheetnames
    search_terms = [tumor_input.strip(), guideline_title.strip()]
    search_terms = [t for t in search_terms if t]

    best_name = None
    best_score = 0.0

    for name in candidates:
        name_low = name.lower()
        for t in search_terms:
            ratio = difflib.SequenceMatcher(a=name_low, b=t.lower()).ratio()
            if ratio > best_score:
                best_score = ratio
                best_name = name

    if best_name and best_score >= 0.55:
        return best_name

    for name in candidates:
        if tumor_input.lower() in name.lower():
            return name
        if guideline_title.lower() in name.lower():
            return name

    return None


# ============================================================
# NCCN Category-1 parsing (unused in your local mode, but kept)
# ============================================================
def parse_category1(html: str) -> List[dict]:
    soup = BeautifulSoup(html, "html.parser")
    items: List[dict] = []

    for block in soup.select("div.item-name"):
        a_tag = block.find("a", href=True)
        if not a_tag:
            continue

        title = a_tag.get_text(strip=True)
        if "Patient" in title:
            continue

        href = a_tag["href"]
        url = BASE_URL + href if href.startswith("/") else href

        version_div = block.find_next_sibling("div", class_="item-version")
        version = "N/A"
        if version_div:
            m = re.search(r"Version[:\s]*([\d.]+)", version_div.get_text())
            if m:
                version = m.group(1)

        items.append({"title": title, "version": version, "url": url})

    seen = set()
    out = []
    for g in items:
        if g["title"] not in seen:
            out.append(g)
            seen.add(g["title"])
    return out

def best_guideline_match(tumor_input: str, guides: List[dict]) -> dict:
    t = tumor_input.strip().lower()

    substring_hits = [g for g in guides if t and t in g["title"].lower()]
    if substring_hits:
        return substring_hits[0]

    best = None
    best_score = 0.0
    for g in guides:
        score = difflib.SequenceMatcher(a=t, b=g["title"].lower()).ratio()
        if score > best_score:
            best_score = score
            best = g

    if best and best_score >= 0.45:
        return best

    raise RuntimeError(f"Could not match tumor input '{tumor_input}' to any NCCN Category 1 guideline title.")

def get_pdf_url_from_detail(detail_url: str) -> str:
    resp = requests.get(detail_url, headers=HEADERS, timeout=120)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")

    link = soup.find("a", href=re.compile(r"/professionals/physician_gls/pdf/.*\.pdf"))
    if not link:
        raise RuntimeError(f"No PDF link found on detail page: {detail_url}")

    href = link["href"]
    return BASE_URL + href if href.startswith("/") else href


# ============================================================
# Selenium download (AUTO login)
# ============================================================
def selenium_download_pdf_with_login(
    nccn_user: str,
    nccn_pass: str,
    pdf_url: str,
    download_dir: Path,
    out_name: str,
    *,
    headless: bool = False,
    download_timeout_s: int = 240,
) -> Path:
    base_url = "https://www.nccn.org"
    login_url = f"{base_url}/login?ReturnURL={pdf_url}"

    download_dir.mkdir(parents=True, exist_ok=True)
    final_pdf = download_dir / out_name
    temp_pdf = download_dir / f"{out_name}.crdownload"

    if final_pdf.exists():
        final_pdf.unlink()
    if temp_pdf.exists():
        temp_pdf.unlink()

    options = webdriver.ChromeOptions()
    if headless:
        options.add_argument("--headless=new")

    options.add_argument(f"--user-data-dir={CHROME_USER_DATA_DIR}")
    options.add_argument(f"--profile-directory={CHROME_PROFILE_DIR}")
    options.add_argument("--no-first-run")
    options.add_argument("--no-default-browser-check")
    options.add_argument("--disable-notifications")
    options.add_argument("--disable-gpu")
    options.add_argument("--disable-features=TranslateUI")
    options.add_argument("--start-maximized")

    prefs = {
        "download.default_directory": str(download_dir.resolve()),
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "plugins.always_open_pdf_externally": True,
    }
    options.add_experimental_option("prefs", prefs)

    driver = webdriver.Chrome(options=options)

    def wait_for_download(timeout_s: int) -> Path:
        end = time.time() + timeout_s
        while time.time() < end:
            if final_pdf.exists() and not temp_pdf.exists():
                return final_pdf
            time.sleep(1)
        raise RuntimeError(f"PDF download failed (timeout after {timeout_s}s).")

    try:
        wait = WebDriverWait(driver, 25)

        # Try direct PDF first (cookies might already be valid)
        driver.get(pdf_url)
        time.sleep(3)
        try:
            return wait_for_download(timeout_s=20)
        except Exception:
            pass

        driver.get(login_url)
        time.sleep(2)

        cur = (driver.current_url or "").lower()
        on_login = ("login" in cur) or ("signin" in cur) or ("sign-in" in cur)

        if on_login:
            email_el = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "input[type='email']")))
            pwd_el = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "input[type='password']")))

            email_el.clear()
            email_el.send_keys(nccn_user)
            pwd_el.clear()
            pwd_el.send_keys(nccn_pass)

            login_btn = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "button[type='submit'], input[type='submit']")))
            login_btn.click()
            time.sleep(3)

        driver.get(pdf_url)
        time.sleep(2)
        return wait_for_download(timeout_s=download_timeout_s)

    finally:
        driver.quit()


# ============================================================
# PDF UPDATES extraction (KEEP STRUCTURE + SAFER STRIKE FILTER)
# ============================================================
def _is_strike_font(fontname: str) -> bool:
    """
    IMPORTANT:
    - Do NOT include 'del' here. Many PDF fonts include 'del' as part of the name
      and you'll accidentally remove non-strikethrough text.
    """
    f = (fontname or "").lower()
    return ("strike" in f) or ("strikethrough" in f) or ("delete" in f)

def _reconstruct_lines_from_words(words: List[dict], y_tol: float = 3.0) -> List[str]:
    if not words:
        return []
    words_sorted = sorted(words, key=lambda w: (round(w["top"] / y_tol) * y_tol, w["x0"]))
    lines: List[str] = []
    current_key = None
    current_words: List[str] = []

    for w in words_sorted:
        key = round(w["top"] / y_tol) * y_tol
        if current_key is None:
            current_key = key
        if key != current_key:
            line = " ".join(current_words).strip()
            if line:
                lines.append(line)
            current_key = key
            current_words = []
        current_words.append(w["text"])

    last = " ".join(current_words).strip()
    if last:
        lines.append(last)

    return lines

def extract_updates_text_and_version(pdf_path: Path) -> Tuple[str, Optional[str]]:
    """
    Goal:
    - Extract UPDATES pages while keeping bullets/headers/newlines.
    - Avoid over-normalization (do NOT flatten to one line).
    """
    updates_pages_text: List[str] = []
    detected_version: Optional[str] = None

    with pdfplumber.open(str(pdf_path)) as pdf:
        for page in pdf.pages:
            raw_text = page.extract_text() or ""
            raw_text = _strip_control_chars(raw_text)
            low = raw_text.lower()

            if detected_version is None:
                m = VERSION_RE.search(raw_text)
                if m:
                    detected_version = m.group(1)

            if "updates in version" not in low:
                continue

            # 1) Best effort: keep visual structure (bullets/headers)
            page_txt = page.extract_text(layout=True, x_tolerance=2, y_tolerance=2) or ""
            page_txt = _strip_control_chars(page_txt)

            # 2) Fallback: words reconstruction (keeps more content if layout fails)
            if not page_txt.strip():
                try:
                    words = page.extract_words(extra_attrs=["fontname", "size"])
                except Exception:
                    words = page.extract_words()

                kept = [w for w in words if not _is_strike_font(w.get("fontname", ""))]
                lines = _reconstruct_lines_from_words(kept, y_tol=2.0)
                page_txt = "\n".join(ln.strip() for ln in lines if ln.strip()).strip()

            # Keep newlines, just trim trailing whitespace; drop noisy lines
            cleaned_lines: List[str] = []
            for line in page_txt.splitlines():
                line = _strip_control_chars(line).rstrip()
                if not line.strip():
                    cleaned_lines.append("")
                    continue
                if _is_noise_line(line):
                    continue
                cleaned_lines.append(line)

            page_txt = "\n".join(cleaned_lines).strip()

            if page_txt:
                updates_pages_text.append(page_txt)

    # Join pages with blank line between pages; keep structure
    txt = "\n\n".join(updates_pages_text)

    # Light cleanup: strip left/right per line, collapse excessive blank lines
    txt = "\n".join(line.strip() for line in txt.splitlines())
    while "\n\n\n" in txt:
        txt = txt.replace("\n\n\n", "\n\n")

    return txt.strip(), detected_version


# ============================================================
# REFERENCE PAYLOADS (global)
# ============================================================
def extract_reference_payloads_from_updates(updates_text: str) -> List[str]:
    payloads: Set[str] = set()

    for m in SECTION_RE.finditer(updates_text):
        sec = m.group(1)
        of_part = m.group(2)
        page_part = m.group(3)

        parts = [sec]
        if of_part:
            parts.append(of_part.replace("  ", " ").strip())
        if page_part:
            parts.append(f"Page {page_part}".strip())

        payloads.add(", ".join(parts))

    for m in MS_RE.finditer(updates_text):
        payloads.add(m.group(0))

    return sorted(payloads)


# ============================================================
# UPDATES -> BLOCKS (THIS FIXES YOUR “MISSING CONTEXT/REFERENCE”)
# ============================================================
def _match_section_header_line(line: str) -> Optional[str]:
    """
    Returns canonical header like: 'GAST-B 3 of 7' if matches, else None.
    """
    s = (line or "").strip()
    m = SECTION_HEADER_LINE_RE.match(s) or SECTION_HEADER_SOFT_RE.match(s)
    if not m:
        return None
    sec = m.group(1).strip()
    a = m.group(2).strip()
    b = m.group(3).strip()
    return f"{sec} {a} of {b}"

def _split_updates_into_sections(lines: List[str]) -> List[Tuple[str, List[str]]]:
    """
    Split the UPDATES text into chunks by NCCN section headers:
    [('GAST-B 3 of 7', [...lines...]), ('GAST-F 2 of 9', [...]), ...]
    If a file has lines before first header, they go under 'UNKNOWN SECTION'.
    """
    out: List[Tuple[str, List[str]]] = []
    current_header = "UNKNOWN SECTION"
    current_lines: List[str] = []

    def flush():
        nonlocal current_lines
        # drop trailing blanks
        while current_lines and not current_lines[-1].strip():
            current_lines.pop()
        if current_lines:
            out.append((current_header, current_lines))
        current_lines = []

    for ln in lines:
        hdr = _match_section_header_line(ln)
        if hdr:
            flush()
            current_header = hdr
            continue
        # keep as-is (bullets preserved)
        current_lines.append(ln)

    flush()
    return out

def _section_block_hits(section_lines: List[str], kws: List[str]) -> bool:
    joined = "\n".join(section_lines).lower()
    return any(kw in joined for kw in kws)

def _tighten_block_around_hits(section_lines: List[str], kws: List[str], pad: int = 4) -> List[str]:
    """
    Extract a smaller window inside a section around the first/last keyword hit,
    while trying to preserve nearby bullets/headers.
    """
    lows = [ln.lower() for ln in section_lines]
    hit_idxs = [i for i, low in enumerate(lows) if any(kw in low for kw in kws)]
    if not hit_idxs:
        return section_lines

    start = max(0, min(hit_idxs) - pad)
    end = min(len(section_lines), max(hit_idxs) + pad + 1)

    # Expand upwards to include immediate bullet parent/topic line if present
    # If previous line is a topic-like line (starts with • / - / ▸), include it.
    while start > 0:
        prev = section_lines[start - 1].strip()
        if not prev:
            start -= 1
            continue
        if prev.startswith(("•", "-", "▸", "▶", "●", "◦")):
            start -= 1
            continue
        break

    # Expand downwards to include continuation lines (often wrapped text)
    while end < len(section_lines):
        nxt = section_lines[end].strip()
        if not nxt:
            break
        # If the next line looks like a continuation (no bullet marker), include 1-2 more
        if not nxt.startswith(("•", "-", "▸", "▶", "●", "◦")):
            end += 1
            continue
        break

    return section_lines[start:end]

def pick_update_blocks_for_question(
    updates_text: str,
    qid: str,
    max_hits: int = 8,
    inner_pad: int = 6,
) -> List[str]:
    """
    NEW behavior:
    - Returns MULTILINE blocks
    - Each block includes its NCCN section header (e.g., GAST-B 3 of 7)
    - Preserves bullets and line breaks (no flattening)
    """
    kws = Q_MAP.get(qid, [])
    if not kws:
        return []

    # Keep original structure; do NOT drop blank lines entirely
    raw_lines = [_strip_control_chars(ln) for ln in updates_text.splitlines()]
    # Filter only the worst noise; keep blank lines
    lines: List[str] = []
    for ln in raw_lines:
        if ln.strip() and _is_noise_line(ln):
            continue
        lines.append(ln.rstrip())

    sections = _split_updates_into_sections(lines)

    out: List[str] = []
    seen: Set[str] = set()

    for (hdr, sec_lines) in sections:
        if not _section_block_hits(sec_lines, kws):
            continue

        excerpt = _tighten_block_around_hits(sec_lines, kws, pad=inner_pad)

        # Compose final block with header on top
        block_lines = [hdr] + [ln for ln in excerpt if ln is not None]

        # Clean up: collapse too many blank lines inside block
        compact: List[str] = []
        blank_run = 0
        for ln in block_lines:
            if not (ln or "").strip():
                blank_run += 1
                if blank_run <= 1:
                    compact.append("")
                continue
            blank_run = 0
            compact.append(ln.rstrip())

        block = "\n".join(compact).strip()

        if block and block not in seen:
            out.append(block)
            seen.add(block)

        if len(out) >= max_hits:
            break

    return out


# ============================================================
# REFERENCES (prefix + version)
# ============================================================
def build_reference_strings(prefix_name: str, version: str, payloads: List[str]) -> List[str]:
    return [f"{prefix_name} NCCN / Version {version} / {p}" for p in payloads]


# ============================================================
# Reporting (TXT)
# ============================================================
def decide_status(existing_answers: List[str], update_blocks: List[str]) -> str:
    if not update_blocks:
        return "NO_SIGNAL"
    if not existing_answers:
        return "NEW_CANDIDATE"
    return "REVIEW"

def _indent_block(block: str, indent: str = "  ") -> str:
    lines = block.splitlines()
    return "\n".join((indent + ln) if ln.strip() else "" for ln in lines)

def _extract_header_ref_from_block(block: str) -> Optional[str]:
    """
    From a block that begins with 'GAST-B 3 of 7', return that header as a reference payload.
    """
    first = (block.splitlines()[:1] or [""])[0].strip()
    hdr = _match_section_header_line(first)
    return hdr

def write_report(
    out_path: Path,
    tumor_input: str,
    guideline_title: str,
    version: str,
    pdf_path: Path,
    sheet_name: Optional[str],
    question_rows_found: List[str],
    per_question: List[Tuple[str, str, List[str], List[str], List[str]]],
    all_refs_full: List[str],
) -> None:
    ensure_dir(out_path.parent)

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    lines: List[str] = []

    lines.append("=" * 90)
    lines.append("NCCN UPDATES VS EXCEL — UPDATE REPORT")
    lines.append("=" * 90)
    lines.append(f"Run time: {now}")
    lines.append(f"Tumor input: {tumor_input}")
    lines.append(f"NCCN matched title: {guideline_title}")
    lines.append(f"NCCN version: {version}")
    lines.append(f"PDF used: {pdf_path}")
    lines.append(f"Excel sheet matched: {sheet_name or 'NOT FOUND'}")
    lines.append(f"Question anchors found in Excel: {', '.join(question_rows_found) if question_rows_found else 'NONE'}")
    lines.append("")

    lines.append("-" * 90)
    lines.append("GLOBAL REFERENCES EXTRACTED FROM UPDATES (prefix preserved)")
    lines.append("-" * 90)
    if all_refs_full:
        for r in all_refs_full:
            lines.append(f"- {normalize_whitespace(r)}")
    else:
        lines.append("(No reference payload tokens extracted from UPDATES pages text.)")
    lines.append("")

    lines.append("=" * 90)
    lines.append("PER-QUESTION REVIEW")
    lines.append("=" * 90)

    for (qid, status, existing_answers, existing_refs, update_blocks) in per_question:
        qtext = next((t for (qq, t) in QUESTION_ANCHORS if qq == qid), "")
        lines.append("")
        lines.append("#" * 90)
        lines.append(f"{qid} — {qtext}")
        lines.append(f"STATUS: {status}")
        lines.append("#" * 90)

        lines.append("EXCEL (existing answer lines):")
        if existing_answers:
            for a in existing_answers[:12]:
                lines.append(f"- {normalize_whitespace(a)}")
            if len(existing_answers) > 12:
                lines.append(f"... ({len(existing_answers) - 12} more lines)")
        else:
            lines.append("(empty / not captured in Excel)")

        lines.append("")
        lines.append("EXCEL (existing references):")
        if existing_refs:
            for r in existing_refs[:12]:
                lines.append(f"- {normalize_whitespace(r)}")
            if len(existing_refs) > 12:
                lines.append(f"... ({len(existing_refs) - 12} more refs)")
        else:
            lines.append("(none)")

        lines.append("")
        lines.append("UPDATES (blocks matched by keywords — with NCCN section header + bullets preserved):")
        if update_blocks:
            for b in update_blocks:
                hdr_ref = _extract_header_ref_from_block(b)
                if hdr_ref:
                    # show a compact reference line first (so your boss can CTRL+F)
                    lines.append(f"- REFERENCE (quick): {hdr_ref}")
                else:
                    lines.append("- REFERENCE (quick): (not detected)")

                # show block indented, multiline, bullets preserved
                lines.append(_indent_block(b, indent="  "))
                lines.append("")  # spacer line between blocks
        else:
            lines.append("(no relevant blocks detected in UPDATES section by current keyword set)")

    out_path.write_text("\n".join(lines), encoding="utf-8")


# ============================================================
# INTERACTIVE HELPERS + CLI RESOLUTION (solid for automation)
# ============================================================
def list_pdfs(pdf_dir: Path) -> List[Path]:
    pdfs = sorted([p for p in pdf_dir.iterdir() if p.is_file() and p.suffix.lower() == ".pdf"])
    return pdfs

def interactive_pick_pdf(pdf_dir: Path) -> Path:
    pdfs = list_pdfs(pdf_dir)
    if not pdfs:
        raise RuntimeError(f"No PDFs found in: {pdf_dir}")

    print("\nAvailable PDFs:\n")
    for i, p in enumerate(pdfs, start=1):
        print(f"{i}. {p.name}")

    while True:
        choice = input("\nType PDF number or name (e.g., 7 or gastric.pdf): ").strip()
        if choice.isdigit():
            idx = int(choice) - 1
            if 0 <= idx < len(pdfs):
                return pdfs[idx]
        for p in pdfs:
            if choice.lower() == p.name.lower() or choice.lower() == p.stem.lower():
                return p
        print("❌ Invalid selection. Try again.")

def resolve_pdf_arg(pdf_arg: str, pdf_dir: Path) -> Path:
    """
    Resolve:
      - full path to .pdf
      - filename in pdf_dir
      - stem without .pdf in pdf_dir
    """
    p = Path(pdf_arg)
    if p.exists() and p.is_file() and p.suffix.lower() == ".pdf":
        return p.resolve()

    candidate = pdf_dir / pdf_arg
    if candidate.exists() and candidate.is_file() and candidate.suffix.lower() == ".pdf":
        return candidate.resolve()

    candidate2 = pdf_dir / f"{pdf_arg}.pdf"
    if candidate2.exists() and candidate2.is_file():
        return candidate2.resolve()

    raise FileNotFoundError(f"Could not find PDF: '{pdf_arg}'. Tried: {p} and {candidate} and {candidate2}")


# ============================================================
# MAIN
# ============================================================
def main():
    parser = argparse.ArgumentParser(
        description="Compare NCCN UPDATES section vs Dropbox Excel (no Excel edits)."
    )

    # Backward compatible positional arg (keeps your current behavior)
    parser.add_argument("pdf", nargs="?", help="PDF filename (e.g., gastric.pdf) OR full path. If omitted, you will pick interactively.")

    # Explicit automation-friendly args (solid for runner/n8n)
    parser.add_argument("--pdf", dest="pdf_flag", default=None, help="Same as positional pdf, but explicit (recommended for automation).")
    parser.add_argument("--pdf-dir", default=str(PDF_GUIDELINES_DIR), help="Directory containing PDFs (default: your PDF Guidelines folder).")
    parser.add_argument("--list-pdfs", action="store_true", help="List available PDFs in --pdf-dir and exit.")
    parser.add_argument("--non-interactive", action="store_true", help="Fail if no --pdf/positional is provided (never prompt).")

    parser.add_argument("--excel", default=str(EXCEL_PATH_DEFAULT), help="Path to the Excel file.")
    parser.add_argument("--out-dir", default=str(OUT_DIR_DEFAULT), help="Folder to save TXT update reports.")
    parser.add_argument("--out", default=None, help="Exact output report path (.txt). If provided, ignores --out-dir naming.")
    parser.add_argument("--tumor", default="", help="Tumor/sheet name override (optional). If omitted, inferred from PDF name.")

    args = parser.parse_args()

    pdf_dir = Path(args.pdf_dir)
    excel_path = Path(args.excel)

    # allow both positional and explicit flag; flag wins if both provided
    pdf_arg = args.pdf_flag or args.pdf

    if args.list_pdfs:
        pdfs = list_pdfs(pdf_dir)
        print("\nAvailable PDFs:\n")
        for i, p in enumerate(pdfs, start=1):
            print(f"{i}. {p.name}")
        return

    if not excel_path.exists():
        raise FileNotFoundError(f"Excel not found: {excel_path}")

    # 1) Choose PDF (flag/positional or interactive)
    if pdf_arg:
        pdf_path = resolve_pdf_arg(pdf_arg, pdf_dir)
    else:
        if args.non_interactive:
            raise RuntimeError("Non-interactive mode: you must provide --pdf <file> (or positional pdf).")
        pdf_path = interactive_pick_pdf(pdf_dir)

    tumor_input = (args.tumor.strip() or pdf_path.stem).strip()
    guideline_title = tumor_input  # local-only mode

    out_dir = Path(args.out_dir)
    ensure_dir(out_dir)

    print("\n==============================")
    print(f"PDF:         {pdf_path}")
    print(f"TUMOR INPUT:  {tumor_input}")
    print("==============================\n")

    # 2) Extract UPDATES + version
    updates_text, detected_version = extract_updates_text_and_version(pdf_path)
    version_for_refs = detected_version or "UNKNOWN"
    print(f"[OK] Detected version from PDF: {version_for_refs}")

    if not updates_text:
        print("[WARN] No UPDATES pages found by heuristic ('Updates in Version'). Report will still be created.")

    payloads = extract_reference_payloads_from_updates(updates_text)
    print(f"[OK] Extracted reference payload tokens from UPDATES: {len(payloads)}")

    # 3) Load Excel and match sheet
    wb = openpyxl.load_workbook(excel_path)
    sheet_name = find_best_sheet(wb, tumor_input=tumor_input, guideline_title=guideline_title)
    ws = wb[sheet_name] if sheet_name else None

    if ws:
        print(f"[OK] Excel sheet matched: {sheet_name}")
        ref_col = find_reference_column(ws)
        q_rows = find_question_rows(ws)
        question_rows_found = sorted(q_rows.keys(), key=lambda x: x)
        blocks = build_blocks(q_rows, ws.max_row)
    else:
        print("[WARN] Could not match an Excel sheet. Report will include UPDATES-only.")
        ref_col = None
        question_rows_found = []
        blocks = []

    prefix_name = sheet_name if sheet_name else tumor_input
    all_refs_full = build_reference_strings(prefix_name, version_for_refs, payloads)

    per_question: List[Tuple[str, str, List[str], List[str], List[str]]] = []

    for qid, _qtxt in QUESTION_ANCHORS:
        # block-based extraction with section header context preserved
        update_blocks = pick_update_blocks_for_question(updates_text, qid=qid, max_hits=8, inner_pad=6)

        if ws and blocks:
            block = next((b for b in blocks if b.qid == qid), None)
            if block:
                existing_answers, existing_refs = extract_existing_block_text(ws, block, ref_col)
            else:
                existing_answers, existing_refs = [], []
        else:
            existing_answers, existing_refs = [], []

        status = decide_status(existing_answers, update_blocks)
        per_question.append((qid, status, existing_answers, existing_refs, update_blocks))

    # Output path: either exact --out, or deterministic name in --out-dir
    if args.out:
        report_path = Path(args.out)
        ensure_dir(report_path.parent)
    else:
        report_name = f"{slugify(prefix_name)}__updates_{version_for_refs}__vs_excel.txt"
        report_path = out_dir / report_name

    write_report(
        out_path=report_path,
        tumor_input=tumor_input,
        guideline_title=guideline_title,
        version=version_for_refs,
        pdf_path=pdf_path,
        sheet_name=sheet_name,
        question_rows_found=question_rows_found,
        per_question=per_question,
        all_refs_full=all_refs_full,
    )

    print(f"\n[DONE] Report written:\n  {report_path}\n")
    print("Excel was NOT modified.")

if __name__ == "__main__":
    main()

